#!/usr/bin/env python3
"""
Restore complete dates from Excel file to SQLite database
"""

import sqlite3
from openpyxl import load_workbook
from datetime import datetime

DATABASE = 'films.db'
EXCEL_FILE = 'original_films.xlsx'

def restore_dates():
    """Restore dates from Excel to database"""

    # Load Excel file
    print(f"Loading {EXCEL_FILE}...")
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    # Connect to database
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    updated = 0
    skipped = 0
    errors = 0

    # Skip header rows (first 2 rows based on the CSV we saw)
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Column structure: empty, Order, Date Film Seen, Film, J-Rayting, Score...
        if not row or len(row) < 4:
            continue

        order_num = row[1]  # Order number
        date_value = row[2]  # Date Film Seen
        title = row[3]      # Film title

        if not title or not order_num:
            continue

        # Convert date to MM/DD/YYYY format if it's a datetime object
        date_str = None
        if isinstance(date_value, datetime):
            date_str = date_value.strftime('%m/%d/%Y')
        elif isinstance(date_value, str):
            # Keep as-is if it's already a string (like "1990s", "April-07", etc.)
            date_str = date_value

        if date_str:
            try:
                # Update the date in the database using order_number to match
                cursor.execute(
                    'UPDATE films SET date_seen = ? WHERE order_number = ?',
                    (date_str, order_num)
                )

                if cursor.rowcount > 0:
                    # Only print if date changed and is an actual date
                    if isinstance(date_value, datetime):
                        print(f"  ✓ Updated #{order_num}: {title} -> {date_str}")
                        updated += 1
                    else:
                        skipped += 1
                else:
                    print(f"  ⚠ No match found for #{order_num}: {title}")
                    errors += 1

            except Exception as e:
                print(f"  ✗ Error updating {title}: {e}")
                errors += 1

    conn.commit()
    conn.close()

    print(f"\n{'='*50}")
    print(f"Summary:")
    print(f"  Dates updated: {updated}")
    print(f"  Skipped (already formatted): {skipped}")
    print(f"  Errors: {errors}")
    print(f"{'='*50}")

if __name__ == '__main__':
    restore_dates()
