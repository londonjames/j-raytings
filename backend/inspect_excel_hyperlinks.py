from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('original_films.xlsx')
ws = wb.active

# Check all cells in first 10 data rows for hyperlinks
print("Inspecting Excel file for hyperlinks...\n")
print("=" * 100)

for row in range(2, 12):  # Check first 10 rows including header
    print(f"\nRow {row}:")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        if cell.hyperlink and cell.hyperlink.target:
            print(f"  Column {col} ({cell.value}): HYPERLINK -> {cell.hyperlink.target}")

print("\n" + "=" * 100)
print("\nChecking if 'Film' column has hyperlinks...")

# Find Film column
headers = [cell.value for cell in ws[2]]
try:
    film_col = headers.index('Film') + 1
except ValueError:
    print("Could not find 'Film' column")
    exit(1)

hyperlink_count = 0
sample_hyperlinks = []

for row in range(3, min(ws.max_row + 1, 50)):  # Check first 50 rows
    cell = ws.cell(row, film_col)
    if cell.hyperlink and cell.hyperlink.target:
        hyperlink_count += 1
        if len(sample_hyperlinks) < 10:
            sample_hyperlinks.append((cell.value, cell.hyperlink.target))

print(f"\nFilm column (column {film_col}) hyperlinks found in first 50 rows: {hyperlink_count}")

if sample_hyperlinks:
    print("\nSample hyperlinks:")
    for title, link in sample_hyperlinks:
        print(f"  {title} -> {link}")
