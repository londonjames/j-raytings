#!/usr/bin/env python3
"""
Extract Notion hyperlinks from Google Sheets Excel export and update database
Usage: Download Google Sheet as Excel (.xlsx), then run this script
"""
import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from import_books import get_db, init_database

def extract_hyperlinks_from_excel(excel_file_path: str) -> dict:
    """
    Extract hyperlinks from Excel export of Google Sheets
    
    Args:
        excel_file_path: Path to the Excel (.xlsx) file exported from Google Sheets
    
    Returns:
        Dictionary mapping order_number to Notion link URL
    """
    notion_links = {}
    
    if not os.path.exists(excel_file_path):
        print(f"❌ Error: Excel file not found: {excel_file_path}")
        return notion_links
    
    if not excel_file_path.endswith(('.xlsx', '.xls')):
        print(f"❌ Error: File must be .xlsx or .xls format")
        return notion_links
    
    print(f"Reading Excel file: {excel_file_path}")
    try:
        wb = load_workbook(excel_file_path, data_only=False)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return notion_links
    
    # Try to find the "all books" sheet
    sheet = None
    for sheet_name in wb.sheetnames:
        if 'book' in sheet_name.lower():
            sheet = wb[sheet_name]
            print(f"Using sheet: {sheet_name}")
            break
    
    if not sheet:
        # Use first sheet if no "books" sheet found
        sheet = wb.active
        print(f"Using active sheet: {sheet.title}")
    
    # Find header row
    header_row_idx = None
    order_col_idx = None
    notion_col_idx = None
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=False), start=1):
        cell_values = [cell.value for cell in row]
        # Look for "Order" column
        if 'Order' in [str(v) for v in cell_values if v]:
            header_row_idx = row_idx
            # Find column indices
            for col_idx, cell in enumerate(row, start=1):
                cell_value = str(cell.value) if cell.value else ''
                if 'Order' in cell_value and order_col_idx is None:
                    order_col_idx = col_idx
                if 'Notes in Notion' in cell_value or ('Notes' in cell_value and 'Notion' in cell_value):
                    notion_col_idx = col_idx
            break
    
    if header_row_idx is None or order_col_idx is None or notion_col_idx is None:
        print("❌ Error: Could not find header row or required columns")
        print(f"  Header row: {header_row_idx}, Order col: {order_col_idx}, Notion col: {notion_col_idx}")
        return notion_links
    
    print(f"Found columns: Order={order_col_idx}, Notes in Notion={notion_col_idx}")
    
    # Also load with data_only=True to get evaluated values (not formulas)
    wb_data = load_workbook(excel_file_path, data_only=True)
    sheet_data = wb_data[sheet.title]
    
    # Load hyperlink relationships from Excel file
    # Excel stores hyperlinks in relationship files, map rId to URLs
    hyperlink_relationships = {}
    try:
        with zipfile.ZipFile(excel_file_path, 'r') as z:
            # Find the relationship file for this sheet
            # Sheet index in openpyxl (0-based) + 1 = sheet number in Excel
            sheet_index = wb.sheetnames.index(sheet.title)
            sheet_num = sheet_index + 1
            rel_file = f'xl/worksheets/_rels/sheet{sheet_num}.xml.rels'
            
            if rel_file in z.namelist():
                content = z.read(rel_file).decode('utf-8')
                root = ET.fromstring(content)
                ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                relationships = root.findall('.//r:Relationship[@Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"]', ns)
                for rel in relationships:
                    rel_id = rel.get('Id')
                    target = rel.get('Target', '')
                    if rel_id and target:
                        hyperlink_relationships[rel_id] = target
                print(f"Loaded {len(hyperlink_relationships)} hyperlink relationships from Excel")
    except Exception as e:
        print(f"Warning: Could not load hyperlink relationships: {e}")
    
    # Process data rows - iterate through actual row numbers
    extracted_count = 0
    yes_count = 0
    
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        # Get order number from formula sheet
        order_cell = sheet.cell(row_idx, order_col_idx)
        order_value = order_cell.value
        if not order_value:
            continue
        
        # Try to get numeric order value (could be formula or number)
        try:
            # If it's a formula like "=B10+1", get the evaluated value from data sheet
            order_cell_data = sheet_data.cell(row_idx, order_col_idx)
            order_value = order_cell_data.value or order_value
            order_number = int(float(str(order_value)))
        except (ValueError, TypeError):
            continue
        
        # Get Notes in Notion cell (from formula sheet for hyperlink object)
        notion_cell = sheet.cell(row_idx, notion_col_idx)
        
        # Get evaluated value from data sheet
        notion_cell_data = sheet_data.cell(row_idx, notion_col_idx)
        notion_value_evaluated = str(notion_cell_data.value).strip().upper() if notion_cell_data.value else ''
        
        # Get formula value from formula sheet
        notion_value_formula = str(notion_cell.value).strip() if notion_cell.value else ''
        notion_value_formula_upper = notion_value_formula.upper()
        
        # Check if it says "YES" (could be direct value or in HYPERLINK formula)
        is_yes = False
        link = None
        
        # Check if evaluated value is YES or formula contains HYPERLINK with YES
        if notion_value_evaluated == 'YES' or ('HYPERLINK' in notion_value_formula_upper and 'YES' in notion_value_formula_upper):
            is_yes = True
        
        if not is_yes:
            continue
        
        yes_count += 1
        
        # Extract link - try multiple methods
        # Method 1: Extract from HYPERLINK formula (most common case)
        if 'HYPERLINK' in notion_value_formula_upper:
            import re
            # Pattern: HYPERLINK("url","text") - extract the URL (first parameter)
            patterns = [
                r'HYPERLINK\s*\(\s*"([^"]+)"',  # Standard: HYPERLINK("url","text")
                r'HYPERLINK\s*\(\s*\'([^\']+)\'',  # Single quotes
            ]
            for pattern in patterns:
                match = re.search(pattern, notion_value_formula, re.IGNORECASE)
                if match:
                    link = match.group(1).strip()
                    link = link.strip('"\'')
                    if link:
                        break
        
        # Method 2: Check hyperlink object (works for direct hyperlinks, not formulas)
        if not link and notion_cell.hyperlink:
            link = notion_cell.hyperlink.target or notion_cell.hyperlink.location
            # If link is a relationship ID (like rId9), look it up in relationships
            if link and (link.startswith('rId') or link.startswith('#')):
                rel_id = link.lstrip('#')
                if rel_id in hyperlink_relationships:
                    link = hyperlink_relationships[rel_id]
                else:
                    link = None
            # Also check hyperlink.id property
            if not link and hasattr(notion_cell.hyperlink, 'id'):
                rel_id = notion_cell.hyperlink.id
                if rel_id in hyperlink_relationships:
                    link = hyperlink_relationships[rel_id]
        
        if link:
            notion_links[order_number] = link
            extracted_count += 1
            if extracted_count <= 5:
                print(f"  ✓ Extracted order {order_number}: {link[:60]}...")
    
    print(f"✓ Found {yes_count} rows with 'YES' in Notes in Notion")
    print(f"✓ Extracted {extracted_count} Notion hyperlinks from Excel")
    return notion_links

def update_database_with_links(notion_links: dict):
    """
    Update database with extracted Notion links
    
    Args:
        notion_links: Dictionary mapping order_number to Notion link URL
    """
    if not notion_links:
        print("No links to update")
        return
    
    print("\nInitializing database...")
    init_database()
    
    db = get_db()
    updated_count = 0
    not_found_count = 0
    already_correct = 0
    
    print(f"\nUpdating {len(notion_links)} Notion links in database...")
    print()
    
    for order_number, notion_link in sorted(notion_links.items()):
        cursor = db.cursor()
        cursor.execute('SELECT id, notion_link FROM books WHERE order_number = ?', (order_number,))
        existing = cursor.fetchone()
        
        if existing:
            book_id, existing_link = existing
            if existing_link != notion_link:
                cursor.execute(
                    'UPDATE books SET notion_link = ? WHERE id = ?',
                    (notion_link, book_id)
                )
                updated_count += 1
                if updated_count <= 10:
                    print(f"  ✓ Updated order {order_number}: {notion_link[:60]}...")
            else:
                already_correct += 1
        else:
            not_found_count += 1
            if not_found_count <= 5:
                print(f"  ⚠ Order {order_number} not found in database")
    
    db.commit()
    db.close()
    
    print()
    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"  Notion links extracted from Excel: {len(notion_links)}")
    print(f"  Updated in database: {updated_count}")
    print(f"  Already up to date: {already_correct}")
    print(f"  Books not found in database: {not_found_count}")
    print()
    print("✓ Notion links update complete!")

def main():
    print("=" * 80)
    print("EXTRACT NOTION LINKS FROM GOOGLE SHEETS EXCEL EXPORT")
    print("=" * 80)
    print()
    print("Instructions:")
    print("1. Open your Google Sheet")
    print("2. File -> Download -> Microsoft Excel (.xlsx)")
    print("3. Run this script with the path to the Excel file")
    print()
    
    if len(sys.argv) < 2:
        print("Usage: python3 extract_notion_links_from_excel.py <path_to_excel_file>")
        print()
        print("Example:")
        print("  python3 extract_notion_links_from_excel.py ~/Downloads/all_books.xlsx")
        return
    
    excel_file = sys.argv[1]
    
    # Extract hyperlinks
    notion_links = extract_hyperlinks_from_excel(excel_file)
    
    if not notion_links:
        print("\n❌ No Notion links extracted. Please check:")
        print("  - Is the Excel file from the correct sheet?")
        print("  - Does it contain the 'Notes in Notion' column?")
        print("  - Are there cells with 'YES' that have hyperlinks?")
        return
    
    # Update database
    update_database_with_links(notion_links)

if __name__ == '__main__':
    main()
