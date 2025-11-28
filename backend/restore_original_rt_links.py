import sqlite3
from openpyxl import load_workbook
from datetime import datetime

# Load the Excel file
print("Loading original Excel file...")
wb = load_workbook('original_films.xlsx')
ws = wb.active

# Find column indices
headers = [cell.value for cell in ws[2]]  # Row 2 has headers
print(f"Headers found: {headers}")

try:
    film_col = headers.index('Film') + 1
    year_col = headers.index('Film Year') + 1
    rt_col = headers.index('Rotten Tomatoes') + 1
except ValueError as e:
    print(f"Error finding columns: {e}")
    exit(1)

print(f"Film column: {film_col}")
print(f"Year column: {year_col}")
print(f"RT column: {rt_col}")

# Extract hyperlinks from RT column
rt_links = {}
print("\nExtracting RT hyperlinks from Excel...")

for row in range(3, ws.max_row + 1):  # Start from row 3 (first data row)
    film_cell = ws.cell(row, film_col)
    year_cell = ws.cell(row, year_col)
    rt_cell = ws.cell(row, rt_col)

    film_title = film_cell.value
    release_year = year_cell.value

    if not film_title:
        continue

    # Check if the RT cell has a hyperlink
    if rt_cell.hyperlink and rt_cell.hyperlink.target:
        rt_link = rt_cell.hyperlink.target
        key = (str(film_title).strip(), release_year)
        rt_links[key] = rt_link

print(f"Found {len(rt_links)} films with RT hyperlinks in Excel")

# Now update the database
conn = sqlite3.connect('films.db')
cursor = conn.cursor()

# Get all films from database
cursor.execute("SELECT id, title, release_year, rt_link FROM films")
db_films = cursor.fetchall()

updated = 0
matched = 0
not_found_in_excel = 0

log_file = f'restore_rt_links_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'

with open(log_file, 'w', encoding='utf-8') as f:
    f.write("=" * 100 + "\n")
    f.write("RESTORING ORIGINAL RT LINKS FROM EXCEL\n")
    f.write("=" * 100 + "\n\n")

    for film_id, title, year, current_rt_link in db_films:
        key = (str(title).strip(), year)

        if key in rt_links:
            excel_link = rt_links[key]

            if current_rt_link != excel_link:
                # Update the database with the ORIGINAL link from Excel
                cursor.execute("""
                    UPDATE films
                    SET rt_link = ?
                    WHERE id = ?
                """, (excel_link, film_id))

                f.write(f"{title} ({year})\n")
                f.write(f"  OLD: {current_rt_link}\n")
                f.write(f"  NEW: {excel_link}\n\n")

                updated += 1
            else:
                matched += 1
        else:
            not_found_in_excel += 1

conn.commit()
conn.close()

print("\n" + "=" * 100)
print("RESTORATION COMPLETE")
print("=" * 100)
print(f"Films with RT links in Excel: {len(rt_links)}")
print(f"Links UPDATED (were wrong): {updated}")
print(f"Links already correct: {matched}")
print(f"Films in DB but no link in Excel: {not_found_in_excel}")
print(f"\nLog saved to: {log_file}")
print("=" * 100)
