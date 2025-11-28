from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('original_films.xlsx')
ws = wb.active

# Find Film column
headers = [cell.value for cell in ws[2]]
film_col = headers.index('Film') + 1

hyperlink_films = []

print("Scanning all rows for RT hyperlinks in Film column...")

for row in range(3, ws.max_row + 1):
    cell = ws.cell(row, film_col)
    if cell.hyperlink and cell.hyperlink.target:
        year_cell = ws.cell(row, headers.index('Film Year') + 1)
        hyperlink_films.append({
            'title': cell.value,
            'year': year_cell.value,
            'link': cell.hyperlink.target,
            'row': row
        })

print(f"\nTotal films with RT hyperlinks: {len(hyperlink_films)}")

# Show samples
print("\nFirst 20 films with hyperlinks:")
for film in hyperlink_films[:20]:
    print(f"  {film['title']} ({film['year']}) -> {film['link']}")

if len(hyperlink_films) > 20:
    print(f"\n... and {len(hyperlink_films) - 20} more")

# Check recent films (2020+)
recent_with_links = [f for f in hyperlink_films if f['year'] and f['year'] >= 2020]
print(f"\nFilms from 2020+ with hyperlinks: {len(recent_with_links)}")

if recent_with_links:
    print("\nRecent films with hyperlinks:")
    for film in recent_with_links[:10]:
        print(f"  {film['title']} ({film['year']}) -> {film['link']}")
